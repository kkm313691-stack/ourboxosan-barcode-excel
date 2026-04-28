import os
import uuid
import datetime
from flask import Flask, request, send_file, jsonify, make_response
from flask_cors import CORS

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.drawing.image import Image

import barcode
from barcode.writer import ImageWriter

app = Flask(__name__)
CORS(app, resources={r"/*": {"origins": "*"}})

@app.after_request
def after_request(response):
    response.headers.add("Access-Control-Allow-Origin", "*")
    response.headers.add("Access-Control-Allow-Headers", "Content-Type")
    response.headers.add("Access-Control-Allow-Methods", "GET, POST, OPTIONS")
    return response


@app.route("/")
def health():
    return jsonify({"status": "ok"})


@app.route("/create_excel", methods=["POST", "OPTIONS"])
def create_excel():
    if request.method == "OPTIONS":
        return make_response("", 200)

    try:
        data = request.json
        mode = data.get("mode", "normal")

        if mode == "lot":
            return create_lot_excel(data)
        else:
            return create_normal_excel(data)

    except Exception as e:
        return jsonify({"error": str(e)}), 500


# =========================
# ✅ 일반 모드 (기존 유지)
# =========================
def create_normal_excel(data):
    name = data.get("name", "")
    exp = data.get("exp", "")
    qty_info = data.get("qty", "")
    qty_generate = int(data.get("barcode_qty", 1))

    today_prefix = datetime.datetime.now().strftime("%Y%m%d")

    wb = Workbook()
    ws = wb.active
    ws.title = "바코드 라벨"

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 140

    label_font = Font(size=40, bold=True)
    value_font = Font(size=100, bold=True)

    align = Alignment(horizontal="center", vertical="center")

    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )

    row = 1
    temp_files = []

    for i in range(1, qty_generate + 1):
        barcode_number = f"{today_prefix}{i:04d}"

        for r in range(row, row + 4):
            ws.row_dimensions[r].height = 200

        labels = ["품명", "소비기한", "수량", "바코드"]
        values = [name, exp, qty_info]

        for idx in range(4):
            a = ws[f"A{row+idx}"]
            b = ws[f"B{row+idx}"]

            a.value = labels[idx]
            a.font = label_font
            a.alignment = align
            a.border = border

            if labels[idx] != "바코드":
                b.value = values[idx]
                b.font = value_font
                b.alignment = align

            b.border = border

        # 바코드 생성 (충돌 방지)
        temp_id = uuid.uuid4().hex
        filename = f"{temp_id}_{i}"
        temp_files.append(filename)

        barcode_class = barcode.get_barcode_class("code128")
        barcode_obj = barcode_class(barcode_number, writer=ImageWriter())
        barcode_obj.save(filename)

        img = Image(f"{filename}.png")
        img.width = 600
        img.height = 150
        ws.add_image(img, f"B{row+3}")

        row += 4

    file_path = f"barcode_{uuid.uuid4().hex}.xlsx"
    wb.save(file_path)

    # 이미지 삭제
    for f in temp_files:
        os.remove(f"{f}.png")

    return send_file(file_path, as_attachment=True, download_name="바코드_라벨.xlsx")


# =========================
# ✅ 로트 모드 (신규)
# =========================
def create_lot_excel(data):
    name = data.get("name", "")
    exp = data.get("exp", "")
    mfg = data.get("mfg", "")
    lot = data.get("lot", "")
    qty_info = data.get("qty", "")
    qty_generate = int(data.get("barcode_qty", 1))

    today_prefix = datetime.datetime.now().strftime("%Y%m%d")

    wb = Workbook()
    ws = wb.active
    ws.title = "로트 바코드"

    ws.column_dimensions["A"].width = 80
    ws.column_dimensions["B"].width = 80

    font_label = Font(size=40, bold=True)
    font_value = Font(size=80, bold=True)

    align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )

    row = 1
    temp_files = []

    for i in range(1, qty_generate + 1):
        barcode_number = f"{today_prefix}{i:04d}"

        for r in range(row, row + 4):
            ws.row_dimensions[r].height = 180

        # 1행
        ws[f"A{row}"].value = "품명"
        ws[f"B{row}"].value = name

        # 2행
        ws[f"A{row+1}"].value = f"소비기한\n{exp}"
        ws[f"B{row+1}"].value = f"제조일자\n{mfg}"

        # 3행
        ws[f"A{row+2}"].value = "수량"
        ws[f"B{row+2}"].value = qty_info

        # 4행
        ws[f"A{row+3}"].value = "바코드"
        ws[f"B{row+3}"].value = lot

        for r in range(row, row + 4):
            for col in ["A", "B"]:
                cell = ws[f"{col}{r}"]
                cell.alignment = align
                cell.border = border

                if r == row+1:
                    cell.font = font_label
                else:
                    cell.font = font_value if col == "B" else font_label

        # 바코드 생성
        temp_id = uuid.uuid4().hex
        filename = f"{temp_id}_{i}"
        temp_files.append(filename)

        barcode_class = barcode.get_barcode_class("code128")
        barcode_obj = barcode_class(barcode_number, writer=ImageWriter())
        barcode_obj.save(filename)

        img = Image(f"{filename}.png")
        img.width = 600
        img.height = 140

        ws.add_image(img, f"A{row+3}")

        # 바코드 번호 표시
        ws[f"A{row+3}"].value = barcode_number

        row += 4

    file_path = f"lot_barcode_{uuid.uuid4().hex}.xlsx"
    wb.save(file_path)

    for f in temp_files:
        os.remove(f"{f}.png")

    return send_file(file_path, as_attachment=True, download_name="로트_바코드.xlsx")
